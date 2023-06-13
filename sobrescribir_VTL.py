from threading import Thread
import openpyxl
import re #libreria para hacer los regex
from tkinter import messagebox as MessageBox
import Configuracion
import jira
import jira_tareas
import jira_usuarios

class SobrescribirVTL(Thread):

    diccionario={}
    diccionarioDescripcion={}
    diccionarioLeido={}
    diccionarioBBDD={}
    filaactual = 1
    valorBase=-1
    wb_VTL = None
    sheet_obj = None
    sheet1_obj = None
    filas = None
    filas1 = None
    multiple_cells = None
    multiple_cells1 = None

    def __init__(self, progressBar):
        super().__init__()
        self.progesBar = progressBar


    """
    list --> boolean
    Esta funcion se utiliza para ver 
    """
    def regex_nohay (self,descrippcion):

        x = re.search("update|Update.packages|package", descrippcion)
        if x:
            parecido = False
        else:
            parecido = True

        return parecido
    """
    lista --> string
    Esta funcion es para una vez se ve que en la lista hay una descripcion parecida a la de parches se saca cual es la descripcion parecida
    """
    def regex_descripcion (self,lista):
        descripcion="none"
        for i in lista:
            x = re.search("update|Update.packages|package", i)
            if x:
                descripcion = i
                break

        return descripcion
    def componente(self,lista,tarea):
        component = lista[3]
        if component in self.diccionarioBBDD.keys():
            self.diccionarioBBDD[component][lista[0]]=[tarea,lista[12],lista[6],lista[7],lista[8],lista[9],lista[10],lista[11],lista[13]]
        else:
            self.diccionarioBBDD[component]={lista[0]:[tarea,lista[12],lista[6],lista[7],lista[8],lista[9],lista[10],lista[11],lista[13]]}
    def escribirCeldas(self,num,lista,hoja):
        solution = 'B' + str(num) 
        nombre='C'+ str(num)
        id = 'D' + str(num)
        hoja[solution]=lista[12] # el numero 12 es la posicion que tiene la descripticon en la lista
        hoja[nombre]= lista[3] # el numero 4 es la posicion que tiene el nombre en la lista
        hoja[id]= lista[0]
    
    def recorrer_Vulnerabilities(self,filaactual):
        for row in self.multiple_cells1:
            fila=[]
            filaactual = filaactual + 1    
            for cell in row:
                fila.append(cell.value)
            
            print (fila)
            if fila[0] in self.diccionarioLeido.keys():
                self.diccionarioLeido[fila[0]]=[fila[1],fila[2],fila[3],fila[4],fila[5],fila[6],filaactual]
            else:
                self.diccionarioLeido[fila[0]]=[fila[1],fila[2],fila[3],fila[4],fila[5],fila[6],filaactual]
    def escrituraHoja1 (self,diccionario,diccionarioLeido,sheet1_obj,filas1):
        num=filas1
        for i in diccionario.keys(): # este for recorre los compoentes que hay en el diccionario
            for x in diccionario[i]: # este for recorre los cve del diccionario del componente anteiror
                valores = diccionario[i][x] # esto es el diccionari, i = al componente , x= cve
                if  valores[0] not in diccionarioLeido.keys():
                    num= num+1
                    task = 'A' + str(num)
                    nombreComponente = 'B' + str(num) 
                    solucion ='C'+ str(num)
                    criticalityScore = 'D' + str(num) 
                    criticalitySeverity = 'E' + str(num) 
                    status = 'F' + str(num) 
                    dateOfDiscovery = 'G' + str(num)
                    correctionsApplied = 'H' + str(num) 
                    installationSteps = 'I' + str(num) 
                    evidences = 'J' + str(num) 
                    notes = 'K' + str(num)  
                    sheet1_obj[solucion]=valores[1]
                    sheet1_obj[nombreComponente]= i
                    print(valores[0]) 
                    sheet1_obj[task]= valores[0]
                    print (valores)
                    sheet1_obj[dateOfDiscovery]= valores[2]
                    listaCriticalidadOut = [] # almacenamos la criticidades mas altas de cada cve de la misma tarea
                    listaEstadoOut = [] # almacenamos el estado de cada cve de la misma tarea
                    for id in diccionario[i]: # recorremos todo el diccionario para ver los cve con las mismas tareas
                        valoresRepetidos = diccionario[i][id] # lista de cada valores de los cve que vamos recorriendo (esto se usa para pillar los valores y el estado de los cve de una misma tarea)
                        if valoresRepetidos[0] == valores [0]: # aqui decimos si es de una misma tarea
                            score1 = -1
                            if valoresRepetidos[7] != "N/A": # esta es la criticidad 3.1 con recalculo
                                score1 = float(valoresRepetidos[7])
                            elif valoresRepetidos[6] != "N/A" and score1== -1: # esta es la criticidad 3.0 realculo
                                score1 = float(valoresRepetidos[6])
                            elif valoresRepetidos[5] != "N/A" and score1== -1: # esta es la criticidad 3.1
                                score1 = float(valoresRepetidos[5])
                            elif valoresRepetidos[4] != "N/A" and score1== -1: # esta es la criticidad 3.0 
                                score1 = float(valoresRepetidos[4])
                            elif valoresRepetidos[3] != "N/A" and score1== -1: # esta es la criticidad 2.0 
                                score1 = float(valoresRepetidos[3])
                            elif score1 == -1:
                                score1 = "N/A"
                            listaCriticalidadOut.append(score1) # añadimos la criticidad mas alta de cada cve a la lista de criticidades
                            listaEstadoOut.append(valoresRepetidos[8]) # añadimos el estado de cada cve a la lista de estados
                    estado = "Close" # Es el estado por defecto solo se cambia si hay alguno abierto
                    for estadoActual in listaEstadoOut: #estadoActual es el estado que se saca en el for de cada cve
                        if estadoActual == "Open":
                            estado = "Open"
                    sheet1_obj[status]=estado # escribimos el estado de la tarea (esta es close solo si todos los cve de la tarea son close)
                    criticidad= -1 # incializamos asi la criticidad ya que la criticidad puede ser 0.0
                    for criticidadActual in listaCriticalidadOut: #calculamos la criciticidad mas alta
                        if float(criticidadActual) > criticidad:
                            criticidad = float(criticidadActual)
                    sheet1_obj[criticalityScore] = criticidad #escribimos en el excel la criticidad mas alta que tenemos 
                    if criticidad == 0.0: # en funcion de la criticidad que tenemos es de un tipo y esto es lo que se calcula
                        severity = "None"
                    elif criticidad > 0.0 and criticidad < 4:
                        severity = "Low"
                    elif criticidad >= 4 and criticidad < 7:
                        severity ="Medium"
                    elif criticidad >= 7 and criticidad < 9:
                        severity ="High"
                    elif criticidad >= 9 and criticidad <= 10:
                        severity ="Critical"
                    sheet1_obj[criticalitySeverity]=severity
                    diccionarioLeido[valores[0]]=[i,valores[1],criticidad,severity,estado,valores[2],num]
                    jira.Jira.conexion(self,valores[0],valores[1])
                    jira_tareas.JiraTareas.ultima_tarea(self)
                    if i == "servidor":
                        Configuracion.Configuracion.component_global= Configuracion.Configuracion.component_servidor
                    else:
                        Configuracion.Configuracion.component_global= Configuracion.Configuracion.component_ACU
                    jira_usuarios.JiraUsuario.asignar_usuario(self)
                
                else:
                    criticalityScore = 'D' + str(diccionarioLeido[valores[0]][6]) 
                    criticalitySeverity = 'E' + str(diccionarioLeido[valores[0]][6]) 
                    status = 'F' + str(diccionarioLeido[valores[0]][6]) 
                    valoresInside = diccionarioLeido[valores[0]]
                    listaCriticalidadInside = [] # almacenamos la criticidades mas altas de cada cve de la misma tarea
                    listaEstadoInside = [] # almacenamos el estado de cada cve de la misma tarea
                    if valoresInside[2] == "N/A":
                            listaCriticalidadInside.append(0)
                    else:
                            v= float(valoresInside[2])
                            listaCriticalidadInside.append(v)
                    for id in diccionario[i]: # recorremos todo el diccionario para ver los cve con las mismas tareas
                        score1 = -1
                        valoresRepetidos = diccionario[i][id] # lista de cada valores de los cve que vamos recorriendo (esto se usa para pillar los valores y el estado de los cve de una misma tarea)
                        if valoresRepetidos[0] == valores [0]: # aqui decimos si es de una misma tarea
                            if valoresRepetidos[7] != "N/A": # esta es la criticidad 3.1 con recalculo
                                score1 = float(valoresRepetidos[7])                           
                            elif valoresRepetidos[6] != "N/A" and score1== -1: # esta es la criticidad 3.0 realculo
                                score1 = float(valoresRepetidos[6])                            
                            elif valoresRepetidos[5] != "N/A" and score1== -1: # esta es la criticidad 3.1
                                score1 = float(valoresRepetidos[5])                            
                            elif valoresRepetidos[4] != "N/A" and score1== -1: # esta es la criticidad 3.0 
                                score1 = float(valoresRepetidos[4])                
                            elif valoresRepetidos[3] != "N/A" and score1== -1: # esta es la criticidad 2.0 
                                score1 = float(valoresRepetidos[3])
                            listaCriticalidadInside.append(score1) # añadimos la criticidad mas alta de cada cve a la lista de criticidades
                            listaEstadoInside.append(valoresRepetidos[8]) # añadimos el estado de cada cve a la lista de estados
                    estado = "Close" # Es el estado por defecto solo se cambia si hay alguno abierto
                    for estadoActual in listaEstadoInside: #estadoActual es el estado que se saca en el for de cada cve
                        if estadoActual == "Open":
                            estado = "Open"
                    sheet1_obj[status]=estado # escribimos el estado de la tarea (esta es close solo si todos los cve de la tarea son close)
                    criticidad= -1 # incializamos asi la criticidad ya que la criticidad puede ser 0.0
                    for criticidadActual in listaCriticalidadInside: #calculamos la criciticidad mas alta
                        if float(criticidadActual) > criticidad:
                            criticidad = float(criticidadActual)
                    sheet1_obj[criticalityScore] = criticidad #escribimos en el excel la criticidad mas alta que tenemos 
                    if criticidad == 0.0: # en funcion de la criticidad que tenemos es de un tipo y esto es lo que se calcula
                        severity = "None"
                    elif criticidad > 0.0 and criticidad < 4:
                        severity = "Low"
                    elif criticidad >= 4 and criticidad < 7:
                        severity ="Medium"
                    elif criticidad >= 7 and criticidad < 9:
                        severity ="High"
                    elif criticidad >= 9 and criticidad <= 10:
                        severity ="Critical"
                    sheet1_obj[criticalitySeverity]=severity
                    diccionarioLeido[valores[0]]=[i,valores[1],criticidad,severity,estado,valores[2],diccionarioLeido[valores[0]][6]]

    def escrituraHoja2(self):
        excel_VL_sobrescribir = Configuracion.Configuracion.excel_VL_sobrescribir
        wb_VL = openpyxl.load_workbook(excel_VL_sobrescribir)
        hojaVL = wb_VL.active
        ultimaFila=(wb_VL.worksheets[0].max_row)
        primeraCelda='A3' 
        ultimaCelda='X' + str(ultimaFila) 
        rango_celdas = hojaVL[primeraCelda:ultimaCelda]
        num=1
        tarea = "TK-TC"+ Configuracion.Configuracion.numero_estacion_sobrescribir +"SEC-"
        numTarea = self.valorBase + 1
        numero = self.filas
        for linea in rango_celdas:
            filaCSR=[]    
            for celda in linea:
                filaCSR.append(celda.value)

            numero = numero + 1
            task= 'A' + str(numero)
            if filaCSR[3] in self.diccionarioDescripcion.keys():
                if filaCSR[0] not in self.diccionarioDescripcion[filaCSR[3]].keys():
                    if  filaCSR[12] not in self.diccionarioDescripcion[filaCSR[3]].keys() :
                        if self.regex_nohay(filaCSR[12]):
                            if numTarea <= 9 : 
                                numeroTarea= "000"+ str(numTarea)
                                nombreTarea = tarea + str(numeroTarea)
                                numTarea = numTarea + 1
                                self.sheet_obj[task]= nombreTarea 
                                self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                            elif numTarea <= 99:
                                numeroTarea= "00" + str(numTarea)
                                nombreTarea = tarea + str(numeroTarea)
                                numTarea = numTarea + 1
                                self.sheet_obj[task]= nombreTarea
                                self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                            elif numTarea <= 999:
                                numeroTarea="0" + str(numTarea)
                                nombreTarea = tarea + str(numeroTarea)
                                numTarea = numTarea + 1
                                self.sheet_obj[task]= nombreTarea
                                self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                            else: 
                                numeroTarea = str(numTarea)
                                nombreTarea = tarea + str(numeroTarea)
                                self.sheet_obj[task]= nombreTarea
                                self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                            self.diccionarioDescripcion[filaCSR[3]][filaCSR[12]]=nombreTarea
                        else:
                            descripcionActualizaciones= self.regex_descripcion(self.diccionarioDescripcion[filaCSR[3]].keys())
                            if descripcionActualizaciones == "none":
                                numeroTarea= "000"+ str(numTarea)
                                nombreTarea = tarea + str(numeroTarea)
                                numTarea = numTarea + 1
                                self.sheet_obj[task] = nombreTarea # esto es escribir la tarea correspondiente en el excel
                                self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                                self.diccionarioDescripcion[filaCSR[3]][filaCSR[12]]=nombreTarea
                            else:
                                nombreTarea = self.diccionarioDescripcion[filaCSR[3]].get(descripcionActualizaciones)# esto saca el valor de la clave del diccionario, que esta se saca con la funcion regex_descripcion
                                self.sheet_obj[task] = nombreTarea # esto es escribir la tarea correspondiente en el excel
                                self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                    else:
                        nombreTarea = self.diccionarioDescripcion[filaCSR[3]].get(filaCSR[12]) # para que te devuelva la tarea que tiene la descripcion que esta repetida, esto da el valor de la clave del diccionario que le pongas
                        self.sheet_obj[task]= nombreTarea
                        self.escribirCeldas(numero,filaCSR,self.sheet_obj)       
                    self.componente(filaCSR,nombreTarea)
                else:
                    numero = numero-1
            else:
                if numTarea <= 9 :
                    numeroTarea= "000"+ str(numTarea)
                    nombreTarea = tarea + str(numeroTarea)
                    numTarea = numTarea + 1
                    self.sheet_obj[task]= nombreTarea
                    self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                elif numTarea <= 99:
                    numeroTarea= "00" + str(numTarea)
                    nombreTarea = tarea + str(numeroTarea)
                    numTarea = numTarea + 1
                    self.sheet_obj[task]= nombreTarea
                    self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                elif numTarea <= 999:
                    numeroTarea="0" + str(numTarea)
                    nombreTarea = tarea + str(numeroTarea)
                    numTarea = numTarea + 1
                    self.sheet_obj[task]= nombreTarea
                    self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                else: 
                    numeroTarea = str(numTarea)
                    nombreTarea = tarea + str(numeroTarea)
                    self.sheet_obj[task]= nombreTarea
                    self.escribirCeldas(numero,filaCSR,self.sheet_obj)
                self.diccionarioDescripcion[filaCSR[3]]={filaCSR[12]:nombreTarea}
                self.componente(filaCSR,nombreTarea)

    def lectura_VTL(self):
        path = Configuracion.Configuracion.nombre_VTL_sobrescrito
        
        self.wb_VTL = openpyxl.load_workbook(path)
        self.sheet_obj = self.wb_VTL.active
        self.sheet1_obj = self.wb_VTL
        self.sheet_obj = self.wb_VTL["Vulnerabilities"]
        self.sheet1_obj= self.wb_VTL["Assessment"]
        
        self.filas = self.sheet_obj.max_row
        self.filas1 = self.sheet1_obj.max_row

        ultimaFila=(self.filas)
        ultimaFila1=(self.filas1)
        primeraCelda='A2'
        ultimaCelda='D' + str(ultimaFila)
        primeraCelda1='A2'
        ultimaCelda1='K' + str(ultimaFila1)
        self.multiple_cells = self.sheet_obj[primeraCelda:ultimaCelda]
        self.multiple_cells1 = self.sheet1_obj[primeraCelda1:ultimaCelda1]
        

        for row in self.multiple_cells:
            fila=[]    
            for cell in row:
                fila.append(cell.value)
            if fila[2] in self.diccionario.keys():
                self.diccionario[fila[2]][fila[3]]=[fila[0],fila[1]]
                self.diccionarioDescripcion[fila[2]][fila[1]]=fila[0]  
            else:
                self.diccionario[fila[2]]={fila[3]:[fila[0],fila[1]]}
                self.diccionarioDescripcion[fila[2]]={fila[1]:fila[0]}
        lista=[]
        for i in range(2, self.filas):
            num1 = self.sheet_obj[f'A{i}']
            valor = num1.value.split(sep="-")
            lista.append(valor[3])
        for x in lista:
            if int(x) > self.valorBase:
                self.valorBase = int(x)
                

    def run(self):
        self.lectura_VTL()
        self.escrituraHoja2()
        self.recorrer_Vulnerabilities(self.filaactual)
        self.escrituraHoja1(self.diccionarioBBDD,self.diccionarioLeido,self.sheet1_obj,self.filas1)
        nombre_VTL_sobrescrito_creado = Configuracion.Configuracion.nombre_VTL_sobrescrito_creado
        ruta = Configuracion.Configuracion.ruta_guardado_sobrescribir
        self.wb_VTL.save(ruta + "/" +nombre_VTL_sobrescrito_creado)
        self.progesBar.set(100)
        MessageBox.showwarning("Alerta", 
        "Sobrescritura del documento VTL terminada.")