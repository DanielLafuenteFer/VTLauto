from openpyxl import load_workbook
import openpyxl
from openpyxl.styles  import PatternFill, Font, Border, Side, Alignment
import Configuracion
import re #libreria para hacer los regex
from tkinter import messagebox as MessageBox
from threading import Thread
import jira
import jira_usuarios
import jira_tareas
class CrearVTL(Thread):

    wb = None
    hoja = None
    hoja1 = None
    borde_hoja_Vulnerabilities = None
    alineamiento_hoja_Vulnerabilities = None
    alineamiento_celdas_insertar = None
    patternFill_hoja_Vulnerabilities = None
    fuente_hoja_Vulnerabilities = None
    borde_hoja_Assessment = None
    alineamiento_hoja_Assessment = None
    patternFill_hoja_Assessment= None
    fuente_hoja_Assessment = None
    listadescripciones={} # diccionario para guardar las descripciones que ya estan en el excel y asi poder ver cuales estan repetidas
    diccionarioBBDD={}
    diccionarioCER={}

    def __init__(self, progressBar):
        super().__init__()
        self.progesBar = progressBar

    """
    list --> boolean
    Esta funcion se utiliza para ver 
    """
    def regex_nohay (self,descrippcion):

        x = re.search("update|Update.packages|package", descrippcion)
        if x:
            parecido = False
        else:
            parecido = True

        return parecido
    """
    lista --> string
    Esta funcion es para una vez se ve que en la lista hay una descripcion parecida a la de parches se saca cual es la descripcion parecida
    """
    def regex_descripcion (self,lista):
        descripcion = "none"
        for i in lista:
            x = re.search("update|Update.packages|package", i)
            if x:
                descripcion = i
                break
        

        return descripcion
    def componente(self,fila,diccionario,tarea): 
        component = fila[3]
        if component in diccionario.keys():
            diccionario[component][fila[0]]=[tarea,fila[12],fila[6],fila[7],fila[8],fila[9],fila[10],fila[11],fila[13]]
        else:
            diccionario[component]={fila[0]:[tarea,fila[12],fila[6],fila[7],fila[8],fila[9],fila[10],fila[11],fila[13]]}
    def escrituraHoja1 (self,diccionario,diccionarioCER,hoja):
        for i in diccionario.keys(): # este for recorre los compoentes que hay en el diccionario
            listaTareasQueEstan = []
            if diccionarioCER!={} or i in diccionarioCER.keys():
                diccionarioCER[i]={}
            else :
                num=1
                for x in diccionario[i]: # este for recorre los cve del diccionario del componente anteiror
                    valores = diccionario[i][x] # esto es el diccionari, i = al componente , x= cve
                    if  valores[0] not in listaTareasQueEstan:
                        listaTareasQueEstan.append(valores[0]) # añadimos a la lista la tarea para que no ponga tareas repetidas
                        num= num+1
                        task = 'A' + str(num)
                        nombreComponente = 'B' + str(num) 
                        solucion ='C'+ str(num)
                        criticalityScore = 'D' + str(num) 
                        criticalitySeverity = 'E' + str(num) 
                        status = 'F' + str(num) 
                        dateOfDiscovery = 'G' + str(num)
                        correctionsApplied = 'H' + str(num) 
                        installationSteps = 'I' + str(num)
                        evidences = 'J' + str(num) 
                        notes = 'K' + str(num)  
                        hoja[solucion]=valores[1]
                        hoja[nombreComponente]= i
                        print(valores[0]) 
                        hoja[task]= valores[0] 
                        hoja[dateOfDiscovery]= valores[2]
                        listaCriticalidad = [] # almacenamos la criticidades mas altas de cada cve de la misma tarea
                        listaEstado = [] # almacenamos el estado de cada cve de la misma tarea
                        for id in diccionario[i]: # recorremos todo el diccionario para ver los cve con las mismas tareas
                            valoresRepetidos = diccionario[i][id] # lista de cada valores de los cve que vamos recorriendo (esto se usa para pillar los valores y el estado de los cve de una misma tarea)
                            if valoresRepetidos[0] == valores [0]: # aqui decimos si es de una misma tarea
                                score1 = -1
                                if valoresRepetidos[7] != "N/A": # esta es la criticidad 3.1 con recalculo
                                    score1 = float(valoresRepetidos[7])
                                elif valoresRepetidos[6] != "N/A" and score1== -1: # esta es la criticidad 3.0 realculo
                                    score1 = float(valoresRepetidos[6])
                                elif valoresRepetidos[5] != "N/A" and score1== -1: # esta es la criticidad 3.1
                                    score1 = float(valoresRepetidos[5])
                                elif valoresRepetidos[4] != "N/A" and score1== -1: # esta es la criticidad 3.0 
                                    score1 = float(valoresRepetidos[4])
                                elif valoresRepetidos[3] != "N/A" and score1== -1: # esta es la criticidad 2.0 
                                    score1 = float(valoresRepetidos[3])
                                elif score1 == -1:
                                    score1 = "N/A"
                                listaCriticalidad.append(score1) # añadimos la criticidad mas alta de cada cve a la lista de criticidades
                                listaEstado.append(valoresRepetidos[8]) # añadimos el estado de cada cve a la lista de estados
                        estado = "Close" # Es el estado por defecto solo se cambia si hay alguno abierto
                        for estadoActual in listaEstado: #estadoActual es el estado que se saca en el for de cada cve
                            if estadoActual == "Open":
                                estado = "Open"
                        hoja[status]=estado # escribimos el estado de la tarea (esta es close solo si todos los cve de la tarea son close)
                        criticidad= -1 # incializamos asi la criticidad ya que la criticidad puede ser 0.0
                        for criticidadActual in listaCriticalidad: #calculamos la criciticidad mas alta
                            if float(criticidadActual) > criticidad:
                                criticidad = float(criticidadActual)
                        hoja[criticalityScore] = criticidad #escribimos en el excel la criticidad mas alta que tenemos 
                        if criticidad == 0.0: # en funcion de la criticidad que tenemos es de un tipo y esto es lo que se calcula
                            severity = "None"
                        elif criticidad > 0.0 and criticidad < 4:
                            severity = "Low"
                        elif criticidad >= 4 and criticidad < 7:
                            severity ="Medium"
                        elif criticidad >= 7 and criticidad < 9:
                            severity ="High"
                        elif criticidad >= 9 and criticidad <= 10:
                            severity ="Critical"
                        hoja[criticalitySeverity]=severity # escribimos en el excel que tipo de criticidad es
                        jira.Jira.conexion(self,valores[0],valores[1])
                        jira_tareas.JiraTareas.ultima_tarea(self)
                        if i == "servidor":
                            Configuracion.Configuracion.component_global= Configuracion.Configuracion.component_servidor
                        else:
                            Configuracion.Configuracion.component_global= Configuracion.Configuracion.component_ACU
                        jira_usuarios.JiraUsuario.asignar_usuario(self)
    def escrituraHoja2(self):
        excel_VL = Configuracion.Configuracion.excel_VL
        base = load_workbook(excel_VL)
        hojaBase = base.active
        ultimaFila=(base.worksheets[0].max_row)
        primeraCelda='A3' 
        ultimaCelda='X' + str(ultimaFila)
        multiple_cells = hojaBase[primeraCelda:ultimaCelda]
        num=1
        tarea = "TK-TC" + Configuracion.Configuracion.numero_estacion_crear + "-SEC-"
        numTarea = 0
        for row in multiple_cells:
            fila=[]    
            for cell in row:
                fila.append(cell.value)
            #print (fila)
            num= num+1
            task= 'A' + str(num)
            solution = 'B' + str(num) 
            nombre='C'+ str(num)
            id = 'D' + str(num)
            self.hoja1[solution]=fila[12] # el numero 12 es la posicion que tiene la descripticon en la lista
            self.hoja1[nombre]= fila[3] # el numero 4 es la posicion que tiene el nombre en la lista
            self.hoja1[id]= fila[0] # el numero 0 es la posicion que tiene el id en la lista
            
            
            if  fila[12] not in self.listadescripciones.keys():# para ver si la descripcion de la vulnerabilidad actual esta ya con anterioridad
                if self.listadescripciones=={} or self.regex_nohay(fila[12]): # este if es para si hay descripciones parecidas que sean la misma o no, en caso de haber una pareciada salta la else si no continua.Lo del diccionario sin nada es para el inicio que da error.
                    if numTarea <= 9 :
                        numeroTarea= "000"+ str(numTarea)
                        nombreTarea = tarea + str(numeroTarea)
                        numTarea = numTarea + 1
                        self.hoja1[task]= nombreTarea 
                    elif numTarea <= 99:
                        numeroTarea= "00" + str(numTarea)
                        nombreTarea = tarea + str(numeroTarea)
                        numTarea = numTarea + 1
                        self. hoja1[task]= nombreTarea
                    elif numTarea <= 999:
                        numeroTarea="0" + str(numTarea)
                        nombreTarea = tarea + str(numeroTarea)
                        numTarea = numTarea + 1
                        self.hoja1[task]= nombreTarea
                    else: 
                        numeroTarea = str(numTarea)
                        nombreTarea = tarea + str(numeroTarea)
                        self.hoja1[task]= nombreTarea
                    self.listadescripciones[fila[12]]=nombreTarea # esto es para que se añada al diccionario la descripcion que no esta como clave y a esa clave se le asigne como valor el numro de tarea que tiene
                else:
                    descripcionActualizaciones= self.regex_descripcion(self.listadescripciones.keys())
                    if descripcionActualizaciones == "none":
                        numeroTarea= "000"+ str(numTarea)
                        nombreTarea = tarea + str(numeroTarea)
                        numTarea = numTarea + 1
                        self.hoja1[task] = nombreTarea # esto es escribir la tarea correspondiente en el excel
                        self.listadescripciones[fila[12]]=nombreTarea
                    else:
                        nombreTarea = self.listadescripciones.get(descripcionActualizaciones)# esto saca el valor de la clave del diccionario, que esta se saca con la funcion regex_descripcion
                        self.hoja1[task] = nombreTarea # esto es escribir la tarea correspondiente en el excel
            else:
                nombreTarea = self.listadescripciones.get(fila[12]) # para que te devuelva la tarea que tiene la descripcion que esta repetida, esto da el valor de la clave del diccionario que le pongas
                self.hoja1[task]= nombreTarea
            self.componente(fila,self.diccionarioBBDD,nombreTarea) # esta es para genrar el diccionario que hace de bbdd

    def inicializar_valores_cabeceras(self):
        self.patternFill_hoja_Vulnerabilities = PatternFill(patternType='solid',start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.fuente_hoja_Vulnerabilities = Font(color="FFFFFF")
        self.alineamiento_hoja_Vulnerabilities = Alignment(
                horizontal='center', 
                vertical='center',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
        self.borde_hoja_Vulnerabilities = Border(
                left=Side(border_style="thin", color='000000'),
                right=Side(border_style="thin", color='000000'),
                top=Side(border_style="thin", color='000000'),
                bottom=Side(border_style="thick", color='000000')
            )
        self.patternFill_hoja_Assessment = PatternFill(patternType='solid',start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.fuente_hoja_Assessment = Font(color="FFFFFF")
        self.alineamiento_hoja_Assessment = Alignment(
                horizontal='center', 
                vertical='center',
                text_rotation=0,
                wrap_text=False,
                shrink_to_fit=False,
                indent=0)
        self.borde_hoja_Assessment = Border(
                left=Side(border_style="thin", color='000000'),
                right=Side(border_style="thin", color='000000'),
                top=Side(border_style="thin", color='000000'),
                bottom=Side(border_style="thick", color='000000')
            )
    def formato_hoja_Assessment(self):
        for indice_columna in range(1, self.hoja.max_column+1):
            celda= self.hoja.cell(row=1, column=indice_columna) #celda de la fila1

            patron = self.patternFill_hoja_Assessment
            fuente = self.fuente_hoja_Assessment
            alineado = self.alineamiento_hoja_Assessment

            celda.fill = patron
            celda.font = fuente
            celda.border = self.borde_hoja_Assessment
            celda.alignment = alineado

            valor = celda.value
            if valor == None: valor = "example" #para dar una longitud predefinida a las columnas que tengan vacia la celda de la cabecera
            
            self.hoja.column_dimensions[celda.column_letter].width = len(valor)+2

    def formato_hoja_Vulnerabilities(self):
        for indice_columna in range(1, self.hoja1.max_column+1):
            celda= self.hoja1.cell(row=1, column=indice_columna) #celda de la fila1

            patron = self.patternFill_hoja_Vulnerabilities
            fuente = self.fuente_hoja_Vulnerabilities
            alineado = self.alineamiento_hoja_Vulnerabilities

            celda.fill = patron
            celda.font = fuente
            celda.border = self.borde_hoja_Vulnerabilities
            celda.alignment = alineado

            valor = celda.value
            if valor == None: valor = "example" #para dar una longitud predefinida a las columnas que tengan vacia la celda de la cabecera
            
            self.hoja.column_dimensions[celda.column_letter].width = len(valor)+2

    def crear_libro_VTL(self):
        self.wb = openpyxl.Workbook()
        self.hoja = self.wb.active
        self.hoja.title = "Assessment"
        self.hoja1 = self.wb.create_sheet("Vulnerabilities")

    def configurar_VTL(self):
        lista=('Task code','Component name','Solution','Criticality score','Criticality severity','Status','Date of discovery','Corrections applied','Installation steps','Evidences','Notes')
        self.hoja.append(lista)
        lista1=('Task code','Solution','Component Name','Vulnerability ID (e.g. CVE or other reference)')
        self.hoja1.append(lista1)
        self.formato_hoja_Assessment()
        self.formato_hoja_Vulnerabilities()
        
    def run(self):
        self.crear_libro_VTL()
        self.inicializar_valores_cabeceras()
        self.configurar_VTL()
        self.escrituraHoja2()
        nombre_VTL_creado = Configuracion.Configuracion.nombre_VTL_creado
        ruta = Configuracion.Configuracion.ruta_guardado_crear
        rutaCompleta =ruta+"/"+nombre_VTL_creado
        self.wb.save(rutaCompleta)
        self.escrituraHoja1(self.diccionarioBBDD,self.diccionarioCER,self.hoja)
        self.wb.save(rutaCompleta)
        self.progesBar.set(100)
        MessageBox.showwarning("Alerta", 
        "Generacion del documento VTL terminado.")